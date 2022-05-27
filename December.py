from openpyxl.styles import Font
import openpyxl
import csv
import openpyxl

############################
#                          #
# １２月受講レポートの自動作成 #
#      December.py         #
#                          #
#                          #
############################

# csvファイルを読み込み
csv_file = open("./December.csv", "r", encoding="utf-8", errors="", newline="" )
f = csv.reader(csv_file, delimiter=",", doublequote=True, lineterminator="\r\n", quotechar='"', skipinitialspace=True)

# date配列の初期化
date = []
# time配列の初期化
time = []
# 行数読み込み
for row in f:
    # 3時間以上の日付と時間の取得
    if row[1] < '3:00':
        continue
    # 1列目(日付)を日付配列に格納
    date.append(row[0])
    # 2列目(時間)を時間配列に格納    
    time.append(row[1])

# 12月度受講レポート資料の読み込み
wb = openpyxl.load_workbook("./【12月教育訓練】受講レポート（氏名）.xlsx")
# ワークシートの選択
ws = wb["受講レポート"]
count=1
# 取得したdate配列の長さ文繰り返し
for i in range(len(date)):
    # 時・分の分解
    hour = time[i].split(':')[0]
    minutes = time[i].split(':')[1]

    # 時間のフォーマット整形
    formattime = str(hour) + '時間' + str(minutes) + '分'

    # 日付記入に対応するセルを取得(6倍ずつ)
    daysheet = ws["B"+str(count*6)]
    # 日付記入に対応するセルを取得(6倍ずつ)    
    timesheet = ws["D"+str(count*6)]
    # 勉強時間を記入に対応するセルを取得(6倍+1ずつ)       
    studytimesheet = ws["B"+str(count*6+1)]
    # 勉強タイトルを記入に対応するセルを取得(6倍+2ずつ)    
    studytitlesheet = ws["B"+str(count*6+2)]
    # 勉強内容を記入に対応するセルを取得(6倍+2ずつ)
    studycontentsheet = ws["B"+str(count*6+4)]
    # 日付の値を入力
    daysheet.value = date[i]
    # 時間の値を入力
    timesheet.value = formattime
    
    count += 1
    # 時間に応じて、記載する内容の変更
    if time[i] == '3:00':
        studytimesheet.value = '9時00分~12時00分'
        studytitlesheet.value = 'Laravel Vue.jsの学習'
        studycontentsheet.value = '業務で使用する知識を学習しております。\nバックエンドとフロントエンドをともに学習しております。'
    elif time[i] == '3:30':
        studytimesheet.value = '13時00分~16時30分'
        studytitlesheet.value = 'Laravel Vue.jsの学習'
        studycontentsheet.value = '業務で使用する知識を学習しております。\nバックエンドとフロントエンドをともに学習しております。'        
    elif time[i] == '4:00':
        studytimesheet.value = '9時00分~14時00分'
        studytitlesheet.value = 'Amazon Connectの学習'
        studycontentsheet.value = '業務で使用する知識を学習しております。\nAWSのサービスとの連携を学習しております'
    elif time[i] == '4:30':
        studytimesheet.value = '9時00分~14時30分'
        studytitlesheet.value = 'Amazon Connectの学習'
        studycontentsheet.value = '業務で使用する知識を学習しております。\nAWSのサービスとの連携を学習しております'    

# fontを指定
font = Font(name='Meiryo UI', size=11)

# 全文字のフォントを統一
for row in ws:
    for cell in row:
        ws[cell.coordinate].font = font
wb.save("../【12月教育訓練】受講レポート（川島健嗣）.xlsx")

wb.close()